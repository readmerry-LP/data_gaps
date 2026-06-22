#!/usr/bin/env python3
"""Data Gaps Dashboard Generator - run weekly after new Master is archived."""
import openpyxl, os, re, json, sys, urllib.request, csv as csvmod
from datetime import datetime
from collections import defaultdict

if sys.platform == 'win32':
    ARCHIVE_FOLDER = r'C:\Users\read.merry\OneDrive - Logistics Plus, Inc\DataGaps Reporting\Archived\Master Reports - 2026'
else:
    ARCHIVE_FOLDER = '/sessions/magical-laughing-turing/mnt/Master Reports - 2026'

MIN_DATE    = '2026-05-01'
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, 'index.html')
TABS        = ['Orders', 'Receipts', 'WW ETAs']
CHARTJS_CDN = 'https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js'

def fetch_chartjs():
    try:
        print("Fetching Chart.js for inline embedding...")
        req = urllib.request.Request(CHARTJS_CDN, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=15) as r:
            content = r.read().decode('utf-8')
        print("  Done (" + str(len(content)) + " chars)")
        return '<script>' + content + '</script>'
    except Exception as e:
        print("  Fetch failed (" + str(e) + "), using CDN link instead.")
        return '<script src="' + CHARTJS_CDN + '"></script>'

def is_gap(v):
    if v is None: return False
    s = str(v).strip()
    return bool(s) and s.lower() != 'good'

def parse_week(fname):
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', fname)
    return (m.group(3) + '-' + m.group(1) + '-' + m.group(2)) if m else None

def display_date(iso):
    return datetime.strptime(iso, '%Y-%m-%d').strftime('%m/%d/%y')

EXCEPTIONS_FILE = os.path.join(SCRIPT_DIR, 'exceptions_log.csv')
exceptions = []
if os.path.exists(EXCEPTIONS_FILE):
    with open(EXCEPTIONS_FILE, newline='', encoding='utf-8') as ef:
        for ex in csvmod.DictReader(ef):
            exceptions.append(ex)
    print("Loaded " + str(len(exceptions)) + " exception(s) from exceptions_log.csv")

files = sorted([f for f in os.listdir(ARCHIVE_FOLDER)
                if f.endswith('.xlsx') and 'MASTER' in f.upper()])
print("Found " + str(len(files)) + " Master reports")

records         = []
docket_week_map = {t: defaultdict(set) for t in TABS}

for fname in files:
    week = parse_week(fname)
    if not week:
        print("  Skipping (no date): " + fname); continue
    if week < MIN_DATE:
        print("  Skipping (before " + MIN_DATE + "): " + fname); continue
    print("  " + fname + " -> " + week)
    wb = openpyxl.load_workbook(os.path.join(ARCHIVE_FOLDER, fname), read_only=True, data_only=True)
    for tab in TABS:
        if tab not in wb.sheetnames: continue
        ws       = wb[tab]
        row_iter = ws.iter_rows(values_only=True)
        hdr      = next(row_iter, None)
        if not hdr: continue
        headers = [str(h).strip() if h else '' for h in hdr]
        gap_cols = ([h for h in headers if h == 'ETA Indicator'] if tab == 'WW ETAs'
                    else [h for h in headers if 'Gap Indicator' in h])
        def col_idx(name, _h=headers):
            return _h.index(name) if name in _h else None
        di = col_idx('Docket ID'); wi = col_idx('Warehouse')
        oi = col_idx('Client') if tab == 'WW ETAs' else col_idx('Org Code')
        for row in row_iter:
            if not row or di is None or not row[di]: continue
            docket_id = str(row[di]).strip()
            whs = str(row[wi]).strip() if wi is not None and row[wi] else ''
            org = str(row[oi]).strip() if oi is not None and row[oi] else ''
            row_has_gap = False
            for gc in gap_cols:
                gci = headers.index(gc)
                if gci < len(row) and is_gap(row[gci]):
                    row_has_gap = True
                    records.append(dict(week=week, tab=tab, docket_id=docket_id,
                                        warehouse=whs, org_code=org,
                                        gap_col=gc, gap_value=str(row[gci]).strip()))
            if row_has_gap:
                docket_week_map[tab][docket_id].add(week)
    wb.close()

print("Total gap events: " + str(len(records)))

all_weeks = sorted(set(r['week'] for r in records))
all_whs   = sorted(set(r['warehouse'] for r in records if r['warehouse']))

weekly = {w: {'Orders': 0, 'Receipts': 0, 'WW ETAs': 0, 'total': 0, 'unique_dockets': 0} for w in all_weeks}
weekly_dockets = {w: set() for w in all_weeks}
for r in records:
    weekly[r['week']][r['tab']] += 1
    weekly[r['week']]['total']  += 1
    weekly_dockets[r['week']].add((r['tab'], r['docket_id']))
for w in all_weeks:
    weekly[w]['unique_dockets'] = len(weekly_dockets[w])

whs_total  = defaultdict(int)
whs_weekly = defaultdict(lambda: defaultdict(int))
for r in records:
    if r['warehouse']:
        whs_total[r['warehouse']] += 1
        whs_weekly[r['warehouse']][r['week']] += 1
whs_ranked = sorted(all_whs, key=lambda w: -whs_total.get(w, 0))

gap_col_tab = defaultdict(lambda: defaultdict(int))
gap_val_counts = defaultdict(int)
for r in records:
    gap_col_tab[r['tab']][r['gap_col']] += 1
    gap_val_counts[r['gap_value']] += 1

new_vs_co = {w: {'new': 0, 'carryover': 0} for w in all_weeks}
for tab in TABS:
    for docket_id, weeks_set in docket_week_map[tab].items():
        sw = sorted(weeks_set)
        for i, w in enumerate(sw):
            if w not in new_vs_co: continue
            wi2 = all_weeks.index(w)
            if i > 0 and sw[i-1] == all_weeks[wi2-1]:
                new_vs_co[w]['carryover'] += 1
            else:
                new_vs_co[w]['new'] += 1

recurring = []
for tab in TABS:
    for docket_id, weeks_set in docket_week_map[tab].items():
        if len(weeks_set) >= 3:
            hits = [r for r in records if r['tab'] == tab and r['docket_id'] == docket_id]
            if hits:
                sw = sorted(weeks_set)
                recurring.append(dict(docket_id=docket_id, tab=tab,
                    warehouse=hits[-1]['warehouse'], org_code=hits[-1]['org_code'],
                    weeks_count=len(weeks_set), first_seen=display_date(sw[0]),
                    last_seen=display_date(sw[-1]), weeks=[display_date(x) for x in sw]))
recurring.sort(key=lambda x: -x['weeks_count'])

docket_gap_cols = defaultdict(set)
for r in records:
    clean = r['gap_col'].replace(' Gap Indicator','').replace(' Indicator','')
    docket_gap_cols[(r['tab'], r['docket_id'])].add(clean)

for entry in recurring:
    entry['gap_types'] = sorted(docket_gap_cols.get((entry['tab'], entry['docket_id']), set()))

recurring_total = len(recurring)

gap_type_recur = defaultdict(lambda: {'total': 0, 'recur': 0})
for tab in TABS:
    for col, count in gap_col_tab[tab].items():
        clean = col.replace(' Gap Indicator','').replace(' Indicator','')
        gap_type_recur[clean]['total'] += count
for tab in TABS:
    for docket_id, weeks_set in docket_week_map[tab].items():
        if len(weeks_set) >= 3:
            for gt in docket_gap_cols.get((tab, docket_id), set()):
                gap_type_recur[gt]['recur'] += 1

# --- Warehouse-client recurring analysis ---
_woi_map = {}
for r in recurring:
    key = (r['warehouse'], r['org_code'])
    if key not in _woi_map:
        _woi_map[key] = {'warehouse': r['warehouse'], 'org_code': r['org_code'],
                         'count': 0, 'chronic': 0, 'max_weeks': 0, 'gap_cnts': defaultdict(int)}
    e = _woi_map[key]
    e['count'] += 1
    if r['weeks_count'] >= 7: e['chronic'] += 1
    if r['weeks_count'] > e['max_weeks']: e['max_weeks'] = r['weeks_count']
    for g in r.get('gap_types', []): e['gap_cnts'][g] += 1

_woi_ranked = sorted(_woi_map.values(), key=lambda x: (-x['chronic'], -x['count']))
_whs_org_insights = []
for item in _woi_ranked[:15]:
    n = item['count']
    sg = sorted(item['gap_cnts'].items(), key=lambda x: -x[1])
    _whs_org_insights.append({
        'warehouse': item['warehouse'],
        'org_code':  item['org_code'],
        'count':     n,
        'chronic':   item['chronic'],
        'max_weeks': item['max_weeks'],
        'gap_types': [{'name': g, 'count': c, 'pct': round(c/n*100)} for g, c in sg[:5]]
    })

_org_recur_whs = {}
for r in recurring:
    org = r['org_code']
    if org not in _org_recur_whs: _org_recur_whs[org] = {}
    whs = r['warehouse']
    if whs not in _org_recur_whs[org]:
        _org_recur_whs[org][whs] = {'count': 0, 'gap_cnts': defaultdict(int)}
    _org_recur_whs[org][whs]['count'] += 1
    for g in r.get('gap_types', []): _org_recur_whs[org][whs]['gap_cnts'][g] += 1

_multi_whs_clients = []
for org, wd in _org_recur_whs.items():
    if len(wd) >= 2:
        tot = sum(v['count'] for v in wd.values())
        ag = defaultdict(int)
        for v in wd.values():
            for g, c in v['gap_cnts'].items(): ag[g] += c
        tg = max(ag.items(), key=lambda x: x[1])[0] if ag else ''
        _multi_whs_clients.append({
            'org_code':        org,
            'warehouses':      sorted(wd.keys()),
            'total_recurring': tot,
            'top_gap':         tg
        })
_multi_whs_clients.sort(key=lambda x: -x['total_recurring'])


org_whs   = defaultdict(lambda: defaultdict(int))
org_total = defaultdict(int)
for r in records:
    if r['org_code'] and r['warehouse']:
        org_whs[r['org_code']][r['warehouse']] += 1
        org_total[r['org_code']] += 1
top_orgs = sorted(org_total, key=lambda o: -org_total[o])[:25]

resolution = {}
for i, w in enumerate(all_weeks):
    if i == 0:
        resolution[w] = {'resolved': 0, 'rate': 0}; continue
    prev   = all_weeks[i-1]
    prev_d = {(t, did) for t in TABS for did, ws2 in docket_week_map[t].items() if prev in ws2}
    curr_d = {(t, did) for t in TABS for did, ws2 in docket_week_map[t].items() if w   in ws2}
    resolved = len(prev_d - curr_d)
    rate = round(resolved / len(prev_d) * 100, 1) if prev_d else 0
    resolution[w] = {'resolved': resolved, 'rate': rate}

whs_gap_types_d  = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
whs_weekly_tab_d = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
whs_weekly_doc_d = defaultdict(lambda: defaultdict(set))
for r in records:
    if r['warehouse']:
        whs_gap_types_d[r['warehouse']][r['tab']][r['gap_col']] += 1
        whs_weekly_tab_d[r['warehouse']][r['week']][r['tab']] += 1
        whs_weekly_doc_d[r['warehouse']][r['week']].add((r['tab'], r['docket_id']))

obs = {}
lw  = all_weeks[-1]
total_all = len(records)

nvc_lw = new_vs_co[lw]
co_total = nvc_lw['new'] + nvc_lw['carryover']
obs['latest_week']     = display_date(lw)
obs['carryover_rate']  = round(nvc_lw['carryover'] / co_total * 100) if co_total else 0
obs['carryover_count'] = nvc_lw['carryover']
obs['new_count']       = nvc_lw['new']
obs['first_week']      = display_date(all_weeks[0])
obs['first_total']     = weekly[all_weeks[0]]['total']
obs['last_total']      = weekly[lw]['total']
obs['volume_pct_change'] = round((obs['last_total'] - obs['first_total']) / obs['first_total'] * 100) if obs['first_total'] else 0
if len(all_weeks) >= 4:
    last4 = [weekly[w]['total'] for w in all_weeks[-4:]]
    mean4 = sum(last4) / 4
    obs['plateau']      = (max(last4) - min(last4)) / mean4 < 0.10 if mean4 else False
    obs['last4_totals'] = last4
    obs['last4_weeks']  = [display_date(w) for w in all_weeks[-4:]]
else:
    obs['plateau'] = False; obs['last4_totals'] = []; obs['last4_weeks'] = []

org_sorted = sorted(org_total.items(), key=lambda x: -x[1])
obs['top_clients']     = [{'org': o, 'count': c, 'pct': round(c / total_all * 100)} for o, c in org_sorted[:3]]
obs['top2_client_pct'] = round(sum(c for _, c in org_sorted[:2]) / total_all * 100) if total_all else 0

gap_col_counts = defaultdict(int)
for r in records:
    gap_col_counts[(r['tab'], r['gap_col'])] += 1
top_gc = max(gap_col_counts.items(), key=lambda x: x[1])
orders_total = sum(weekly[w]['Orders'] for w in all_weeks)
obs['top_gap_col']       = top_gc[0][1].replace(' Gap Indicator','').replace(' Indicator','')
obs['top_gap_col_tab']   = top_gc[0][0]
obs['top_gap_col_count'] = top_gc[1]
obs['top_gap_col_pct']   = round(top_gc[1] / orders_total * 100) if orders_total else 0
obs['top_gap_values']    = [{'value': v, 'count': c} for v, c in sorted(gap_val_counts.items(), key=lambda x: -x[1])[:8]]

whs_sorted = sorted(whs_total.items(), key=lambda x: -x[1])
obs['top_warehouses'] = [{'whs': w, 'count': c, 'pct': round(c / total_all * 100)} for w, c in whs_sorted[:5]]
obs['top3_whs_pct']   = round(sum(c for _, c in whs_sorted[:3]) / total_all * 100) if total_all else 0

max_weeks = len(all_weeks)
stubborn = []
for tab in TABS:
    for did, ws_set in docket_week_map[tab].items():
        if len(ws_set) >= max_weeks - 1:
            hits = [r for r in records if r['tab'] == tab and r['docket_id'] == did]
            if hits:
                stubborn.append({'docket_id': did, 'tab': tab,
                    'warehouse': hits[-1]['warehouse'], 'org_code': hits[-1]['org_code'],
                    'weeks': len(ws_set), 'max_weeks': max_weeks})
stubborn.sort(key=lambda x: -x['weeks'])
obs['stubborn_dockets'] = stubborn[:15]
obs['stubborn_count']   = len(stubborn)
obs['max_weeks']        = max_weeks

half = len(all_weeks) // 2
first_half = set(all_weeks[:half]); second_half = set(all_weeks[half:])
whs_f = defaultdict(int); whs_s = defaultdict(int)
for r in records:
    if r['warehouse']:
        if r['week'] in first_half:  whs_f[r['warehouse']] += 1
        if r['week'] in second_half: whs_s[r['warehouse']] += 1
changes = []
for w in set(list(whs_f.keys()) + list(whs_s.keys())):
    f = whs_f.get(w, 0); s = whs_s.get(w, 0)
    if f >= 50:
        changes.append({'whs': w, 'first': f, 'second': s, 'pct': round((s-f)/f*100)})
obs['best_improvers']  = sorted(changes, key=lambda x:  x['pct'])[:3]
obs['worst_degraders'] = sorted(changes, key=lambda x: -x['pct'])[:3]

res_rates = [resolution[w]['rate'] for w in all_weeks[1:] if resolution[w]['rate'] > 0]
obs['avg_resolution_rate']    = round(sum(res_rates)/len(res_rates), 1) if res_rates else 0
obs['latest_resolution_rate'] = resolution[lw]['rate']

top2_names = ' and '.join(c['org'] for c in obs['top_clients'][:2])
top3_whs   = ', '.join(w['whs'] for w in obs['top_warehouses'][:3])
improvers  = [w['whs'] for w in obs['best_improvers'] if w['pct'] < -50]
imp_str    = ' and '.join(improvers[:2]) if improvers else ''

obs['bottom_line'] = (
    'A ' + str(obs['carryover_rate']) + '% week-over-week repeat rate confirms most teams are not acting on the report. '
    + 'Two clients (' + top2_names + ') and three warehouses (' + top3_whs + ') drive the majority of total volume. '
    + 'Targeting them directly would have the largest impact. '
    + obs['top_gap_values'][0]['value'] + ' is the single most fixable issue -- '
    + str(obs['top_gap_col_count']) + ' occurrences that require someone to upload a document in CW. '
    + (('The improvement at ' + imp_str + ' proves the system works when warehouses engage.') if imp_str else '')
)

lw_docket_whs2 = {}
for r in records:
    if r['week'] == lw:
        lw_docket_whs2[(r['tab'], r['docket_id'])] = r['warehouse']

def streak_len(ws_set):
    s = 0
    for w2 in reversed(all_weeks):
        if w2 in ws_set: s += 1
        else: break
    return s

def age_bucket(s):
    if s == 1: return 'new'
    elif s <= 3: return 'short'
    elif s <= 6: return 'medium'
    return 'chronic'

gap_age_breakdown = {'new': 0, 'short': 0, 'medium': 0, 'chronic': 0}
whs_gap_age = {whs: {'new': 0, 'short': 0, 'medium': 0, 'chronic': 0} for whs in whs_ranked}
for tab in TABS:
    for did, ws_set in docket_week_map[tab].items():
        if lw not in ws_set: continue
        s = streak_len(ws_set)
        b = age_bucket(s)
        gap_age_breakdown[b] += 1
        w2 = lw_docket_whs2.get((tab, did), '')
        if w2 in whs_gap_age:
            whs_gap_age[w2][b] += 1

_gtr_sorted     = sorted(gap_type_recur.items(), key=lambda x: -x[1]['total'])
_gap_type_recur = dict()
for k, v in _gtr_sorted:
    _gap_type_recur[k] = {'total': v['total'], 'recur': v['recur']}

_whs_weekly = dict()
for whs in whs_ranked:
    _whs_weekly[whs] = dict()
    for w in all_weeks:
        _whs_weekly[whs][w] = whs_weekly[whs].get(w, 0)

_gap_col_tab = dict()
for tab in TABS:
    _gap_col_tab[tab] = dict(gap_col_tab[tab])

_org_whs   = dict()
_org_total = dict()
for o in top_orgs:
    _org_whs[o]   = dict(org_whs[o])
    _org_total[o] = org_total[o]

_whs_gap_types = dict()
for whs in whs_ranked:
    _whs_gap_types[whs] = dict()
    for tab in TABS:
        _whs_gap_types[whs][tab] = dict(whs_gap_types_d[whs].get(tab, {}))

_whs_doc_count = dict()
for whs in whs_ranked:
    _whs_doc_count[whs] = dict()
    for w in all_weeks:
        _whs_doc_count[whs][w] = len(whs_weekly_doc_d[whs].get(w, set()))

_whs_weekly_tab = dict()
for whs in whs_ranked:
    _whs_weekly_tab[whs] = dict()
    for w in all_weeks:
        _whs_weekly_tab[whs][w] = dict()
        for tab in TABS:
            _whs_weekly_tab[whs][w][tab] = whs_weekly_tab_d[whs].get(w, {}).get(tab, 0)

_exceptions_list = []
for ex in exceptions:
    row = dict()
    row['date_reported']  = ex.get('Date Reported', '')
    row['effective_date'] = ex.get('Effective Date', '')
    row['warehouse']      = ex.get('Warehouse', '')
    row['org_code']       = ex.get('Org Code', '')
    row['gap_type']       = ex.get('Gap Type', '')
    row['carrier_filter'] = ex.get('Carrier Filter', '')
    row['reason']         = ex.get('Reason', '')
    row['reported_by']    = ex.get('Reported By', '')
    row['notes']          = ex.get('Notes', '')
    _exceptions_list.append(row)

data = dict(
    generated      = datetime.now().strftime('%Y-%m-%d %H:%M'),
    file_count     = sum(1 for f in files if (parse_week(f) or '') >= MIN_DATE),
    total_events   = len(records),
    weeks          = all_weeks,
    display_weeks  = [display_date(w) for w in all_weeks],
    warehouses     = whs_ranked,
    weekly         = weekly,
    whs_total      = dict(whs_total),
    whs_weekly     = _whs_weekly,
    gap_col_tab    = _gap_col_tab,
    new_vs_co      = new_vs_co,
    recurring      = recurring[:150],
    recurring_total     = recurring_total,
    whs_org_insights  = _whs_org_insights,
    multi_whs_clients = _multi_whs_clients,
    gap_type_recur = _gap_type_recur,
    top_orgs       = top_orgs,
    org_whs        = _org_whs,
    org_total      = _org_total,
    resolution     = resolution,
    gap_age        = gap_age_breakdown,
    whs_age        = whs_gap_age,
    whs_gap_types  = _whs_gap_types,
    whs_weekly_tab = _whs_weekly_tab,
    whs_doc_count  = _whs_doc_count,
    observations   = obs,
    exceptions     = _exceptions_list,
)

chartjs_tag = fetch_chartjs()
CDN_TAG = '<script src="' + CHARTJS_CDN + '"></script>'
before = open(os.path.join(SCRIPT_DIR, 'template_before.html'), encoding='utf-8').read().replace(CDN_TAG, chartjs_tag)
after  = open(os.path.join(SCRIPT_DIR, 'template_after.html'),  encoding='utf-8').read()

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write(before)
    f.write(json.dumps(data, default=str))
    f.write(after)

print("Dashboard written to: " + OUTPUT_FILE)
